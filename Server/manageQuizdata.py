from openpyxl import load_workbook
import datetime
import datetime
import json
import locale
import pathlib
import random
import sys

locale.setlocale(locale.LC_ALL, "hu_HU.utf8")

sys.path.insert(1, str(pathlib.Path("./modules").resolve()))
import quizDB

cwd = pathlib.Path(__file__).parent.resolve()
DBRoot = cwd / "data"
masterRoot = cwd / "masterdata"
quizDB = quizDB.QuizDB(DBRoot)


txtHeaders = {
    "id": "id",
    "Doboz": "box",
    "Szám": "answer",
    "Magyar Név": "name_hu",
    "Angol Név": "name_en",
    "Magyar Lokáció": "location_hu",
    "Angol Lokáció": "location_en",
    "Tipus": "type",
}

jsonHeaders = ["id", "box", "answer", "name_hu", "name_en", "location_hu", "location_en", "type"]


def JSON_to_DB():
    with open(masterRoot / "masterList.json", "r", encoding="utf-8") as f:
        rawQuizData: dict[str, list[dict[str, str | int]]] = json.load(f)
    if dt := rawQuizData.get("lastEdited"):
        print(f"Master list last edited: {dt}")
    quizData = rawQuizData.get("entries")
    if not quizData:
        raise ValueError(f"Failed to load masterList.json: {rawQuizData}")
    quizDB.cursor.execute("DELETE FROM buildings;")
    quizDB.cursor.executemany(
        f"INSERT INTO buildings ({', '.join(jsonHeaders)}) \
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
        [tuple(entry[k] for k in jsonHeaders) for entry in quizData],
    )
    quizDB.connection.commit()
    print(f"Loaded {quizDB.cursor.rowcount} entries into the database")


def DB_to_JSON():
    quizDB.cursor.execute(f"SELECT {', '.join(jsonHeaders)} FROM buildings ORDER BY name_hu;")
    data = [dict(zip(jsonHeaders, row)) for row in quizDB.cursor.fetchall()]
    jsonData = {"lastEdited": datetime.datetime.now().isoformat(timespec="milliseconds"), "entries": data}
    with open(masterRoot / "masterList.json", "w", encoding="utf-8") as f:
        json.dump(jsonData, f, indent=4, ensure_ascii=False)
    print(f"Saved {len(data)} entries to masterList.json")


def DB_to_XLSX():
    def getBuildingsDataSortedBy(order: str) -> list[dict[str, str | int | None]]:
        return [dict(zip(jsonHeaders, row)) for row in quizDB.cursor.execute(f"SELECT {','.join(jsonHeaders)} FROM buildings ORDER BY ?;", (order,)).fetchall()]

    def getQuizDataSortedBy(order: str, quiz_round: int) -> list[dict[str, str | int | None]]:
        return [
            dict(zip(jsonHeaders, row))
            for row in quizDB.cursor.execute(
                f"SELECT {','.join([f"buildings.{v}" for v in jsonHeaders])} FROM buildings JOIN quizzes ON buildings.id = quizzes.building_id \
                    WHERE quizzes.quiz_round = ? \
                    ORDER BY buildings.{order};",
                (quiz_round,),
            ).fetchall()
        ]

    resultsDir = masterRoot / "ManualTests"
    templateDir = masterRoot / "ManualTests" / "Minták"
    if not resultsDir.exists():
        raise FileNotFoundError(f"Directory {resultsDir} not found")
    if not templateDir.exists():
        raise FileNotFoundError(f"Directory {templateDir} not found")

    print("Generating master lists...")
    wb_lists = load_workbook(templateDir / "MintaListák.xlsx", read_only=False)
    for colindex, val in enumerate(txtHeaders.keys()):
        wb_lists.active.cell(2, colindex + 1, val)
    ws_name = wb_lists.copy_worksheet(wb_lists.active)
    ws_id = wb_lists.copy_worksheet(wb_lists.active)
    ws_box = wb_lists.copy_worksheet(wb_lists.active)
    ws_answer = wb_lists.copy_worksheet(wb_lists.active)
    wb_lists.remove(wb_lists.active)
    ws_name.title = "Név"
    ws_id.title = "ID"
    ws_box.title = "Doboz"
    ws_answer.title = "Válasz"
    # sorting by name_hu
    for rowindex, row in enumerate(getBuildingsDataSortedBy("name_hu")):
        for c, d in enumerate(txtHeaders.values()):
            ws_name.cell(rowindex + 3, c + 1, row[d] or "None")
    # sorting by id
    for rowindex, row in enumerate(getBuildingsDataSortedBy("id")):
        for c, d in enumerate(txtHeaders.values()):
            ws_id.cell(rowindex + 3, c + 1, row[d] or "None")
    # sorting by box
    for rowindex, row in enumerate(getBuildingsDataSortedBy("box")):
        for c, d in enumerate(txtHeaders.values()):
            ws_box.cell(rowindex + 3, c + 1, row[d] or "None")
    # sorting by answer
    for rowindex, row in enumerate(getBuildingsDataSortedBy("answer")):
        for c, d in enumerate(txtHeaders.values()):
            ws_answer.cell(rowindex + 3, c + 1, row[d] or "None")
    wb_lists.save(resultsDir / "Listák.xlsx")

    print("Generating test size 100...")
    quizData = getBuildingsDataSortedBy("name_hu")
    quarterpoint = int((len(quizData) + 3) / 4)
    print("Quarterpoint:", quarterpoint)
    wb_test100 = load_workbook(templateDir / "Minta100asTeszt.xlsx", read_only=False)
    ws_test100_hu = wb_test100["Minta100asTeszt HU"]
    ws_test100_en = wb_test100["Minta100asTeszt EN"]
    ws_test100sols = wb_test100["MintaMegoldókulcs"]
    ws_test100_hu.title = "100-as Teszt HU"
    ws_test100_en.title = "100-as Teszt EN"
    ws_test100sols.title = "100-as Megoldókulcsok"
    ws_test100_hu["A1"] = "100-as Teszt"
    ws_test100_hu["A2"] = f"Kutatók éjszakája {datetime.datetime.now().year}"
    ws_test100_hu["D1"] = "100-as Teszt"
    ws_test100_hu["D2"] = f"Kutatók éjszakája {datetime.datetime.now().year}"
    ws_test100_hu["G1"] = "100-as Teszt"
    ws_test100_hu["G2"] = f"Kutatók éjszakája {datetime.datetime.now().year}"
    ws_test100_hu["J1"] = "100-as Teszt"
    ws_test100_hu["J2"] = f"Kutatók éjszakája {datetime.datetime.now().year}"
    ws_test100_en["A1"] = "Test 100"
    ws_test100_en["A2"] = f"Researchers' Night {datetime.datetime.now().year}"
    ws_test100_en["D1"] = "Test 100"
    ws_test100_en["D2"] = f"Researchers' Night {datetime.datetime.now().year}"
    ws_test100_en["G1"] = "Test 100"
    ws_test100_en["G2"] = f"Researchers' Night {datetime.datetime.now().year}"
    ws_test100_en["J1"] = "Test 100"
    ws_test100_en["J2"] = f"Researchers' Night {datetime.datetime.now().year}"

    for rowindex, row in enumerate(getBuildingsDataSortedBy("name_hu")):
        ws_test100_hu.cell((rowindex % quarterpoint) + 6, int(rowindex / quarterpoint) * 3 + 1, row["name_hu"])
        ws_test100_hu.cell((rowindex % quarterpoint) + 6, int(rowindex / quarterpoint) * 3 + 2, row["location_hu"])
        ws_test100sols.cell((rowindex % quarterpoint) + 3, int(rowindex / quarterpoint) + 1, row["answer"])
    for rowindex, row in enumerate(getBuildingsDataSortedBy("name_en")):
        ws_test100_en.cell((rowindex % quarterpoint) + 6, int(rowindex / quarterpoint) * 3 + 1, row["name_en"])
        ws_test100_en.cell((rowindex % quarterpoint) + 6, int(rowindex / quarterpoint) * 3 + 2, row["location_en"])
        ws_test100sols.cell((rowindex % quarterpoint) + 3, int(rowindex / quarterpoint) + 5, row["answer"])
    wb_test100.save(resultsDir / "100asTeszt.xlsx")

    print("Generating tests size 20...")
    wbTest = load_workbook(templateDir / "MintaTesztek.xlsx", read_only=False)
    wbSol = load_workbook(templateDir / "MintaMegoldókulcsok.xlsx", read_only=False)
    wsTest_hu = wbTest["MintaHU"]
    wsTest_en = wbTest["MintaEN"]
    wsSol_hu = wbSol.active
    wsSol_en = wbSol.copy_worksheet(wsSol_hu)
    wsSol_hu.title = "Kulcs HU"
    wsSol_en.title = "Kulcs EN"
    testnums: list[int] = [r[0] for r in quizDB.cursor.execute("SELECT quiz_round FROM quizzes WHERE quiz_round > 0 GROUP BY quiz_round;").fetchall()]
    for testnum in testnums:
        print(f"Generating test {testnum}...")
        hu = wbTest.copy_worksheet(wsTest_hu)
        en = wbTest.copy_worksheet(wsTest_en)
        hu.title = f"Teszt {testnum:02} HU"
        en.title = f"Teszt {testnum:02} EN"
        hu["A1"] = f"Teszt {testnum:02}"
        en["A1"] = f"Test {testnum:02}"
        hu["A2"] = f"Kutatók éjszakája - {datetime.datetime.now().year}"
        en["A2"] = f"Researchers' Night - {datetime.datetime.now().year}"
        wsSol_hu.cell(1, testnum, f"Teszt {testnum:02} HU")
        wsSol_en.cell(1, testnum, f"Teszt {testnum:02} EN")
        wsSol_hu.cell(22, testnum, f"Teszt {testnum:02} HU")
        wsSol_en.cell(22, testnum, f"Teszt {testnum:02} EN")
        for rowindex, row in enumerate(getQuizDataSortedBy("name_hu", testnum)):
            hu.cell(rowindex + 6, 1, row["name_hu"])
            hu.cell(rowindex + 6, 2, row["location_hu"])
            wsSol_hu.cell(rowindex + 2, testnum, row["answer"])
        for rowindex, row in enumerate(getQuizDataSortedBy("name_en", testnum)):
            en.cell(rowindex + 6, 1, row["name_en"])
            en.cell(rowindex + 6, 2, row["location_en"])
            wsSol_en.cell(rowindex + 2, testnum, row["answer"])
    wbTest.remove(wsTest_hu)
    wbTest.remove(wsTest_en)
    wbTest.save(resultsDir / "Tesztek.xlsx")
    wbSol.save(resultsDir / "Megoldókulcsok.xlsx")


def TXT_to_JSON():
    def format(key: str, value: str) -> str | int | None:
        if key in ["id", "box", "answer"]:
            if value == "None":
                return None
            return int(value)
        else:
            return value

    with open(masterRoot / "MesterLista.txt", "r", encoding="utf-16") as f:
        fileLines = [line.strip().split("\t") for line in f]
    header = fileLines[0]
    lines = fileLines[1:]
    if header != list(txtHeaders.keys()):
        raise ValueError(f"CSV header mismatch: {header} != {txtHeaders}")
    data = [{v: line[i] for i, v in enumerate(txtHeaders.values())} for line in lines]
    jsonData = {
        "lastEdited": datetime.datetime.now().isoformat(timespec="milliseconds"),
        "entries": [{k: format(k, line[k]) for k in jsonHeaders} for line in data],
    }
    with open(masterRoot / "masterList.json", "w", encoding="utf-8") as f:
        json.dump(jsonData, f, indent=4, ensure_ascii=False)
    print(f"Saved {len(data)} entries to masterList.json")


def regenerateQuizzes(quizCount: int = 8, questionsCount: int = 20):
    res = quizDB.cursor.execute("SELECT id FROM buildings;").fetchall()
    availableBuildings = [row[0] for row in res]
    if len(availableBuildings) < questionsCount:
        raise RuntimeError(f"Too few questions in the database ({len(availableBuildings)}) to generate {questionsCount} questions")
    quizzes = [random.sample(availableBuildings, questionsCount) for _ in range(quizCount)]
    quizDB.cursor.execute("DELETE FROM quizzes;")
    quizDB.cursor.execute("DELETE FROM teams;")
    quizDB.cursor.execute("DELETE FROM answers;")
    quizDB.cursor.executemany("INSERT INTO quizzes (quiz_round, building_id) VALUES (?, ?);", [(-1, b) for b in availableBuildings])
    print(f"Generated the 'SIZE_100' quiz with {quizDB.cursor.rowcount} questions")
    for quizNum, quiz in enumerate(quizzes):
        quizDB.cursor.executemany("INSERT INTO quizzes (quiz_round, building_id) VALUES (?, ?);", [(quizNum + 1, b) for b in quiz])
    print(f"Generated {quizCount} quizzes with {quizDB.cursor.rowcount} questions each (taken from the last quiz)")
    quizDB.connection.commit()


def generateAnswers():
    if input("Are you sure? (YES)> ") != "YES":
        return
    res: list[tuple[int | str]] = quizDB.cursor.execute("SELECT id FROM buildings;").fetchall()
    availableBuildings: list[int] = [row[0] for row in res]
    availableAnswers = list(range(1, len(availableBuildings) + 1))
    random.shuffle(availableBuildings)
    random.shuffle(availableAnswers)
    newAnswers = list(zip(availableBuildings, availableAnswers))
    for id in availableBuildings:
        quizDB.cursor.execute("UPDATE buildings SET answer=? WHERE id=?;", (None, id))
    for id, ans in newAnswers:
        quizDB.cursor.execute("UPDATE buildings SET answer=? WHERE id=?;", (ans, id))
    print(f"Generated {len(newAnswers)} answers")
    quizDB.connection.commit()


# ------- MAIN -------
def main(arg: str = ""):
    if arg == "j2d":
        print("Converting JSON to database...")
        JSON_to_DB()
    elif arg == "d2j":
        print("Converting database to JSON...")
        DB_to_JSON()
    elif arg == "d2x":
        print("Converting JSON to XLSX...")
        DB_to_XLSX()
    elif arg == "d2j2x":
        print("Converting database to JSON and then to XLSX...")
        DB_to_JSON()
        DB_to_XLSX()
    elif arg == "regen":
        print("WARNING: Automatically generating quizzes, uncomment line to enable asking")
        # if input("Are you sure to regenerate all quizzes? (YES/NO) > ") == "YES":
        if True:
            quizCount = len(sys.argv) > 2 and int(sys.argv[2]) or 8
            questionCount = len(sys.argv) > 3 and int(sys.argv[3]) or 20
            print(f"Regenerating quizzes with {quizCount} quizzes and {questionCount} questions...")
            regenerateQuizzes(quizCount, questionCount)
    elif arg == "newanswers":
        print("Generating new answers...")
        generateAnswers()
    else:
        # print help
        print(
            "\n".join(
                [
                    "\nUsage: python manageQuizdata.py <task>",
                    "",
                    "<task>:",
                    "- j2d: Convert JSON to database",
                    "- d2j: Convert database to JSON",
                    "- d2x: Convert database to XLSX",
                    "- d2j2x: Convert database to JSON and then to XLSX",
                    "- regen [quizCount] [questionsCount]: Regenerate quizzes",
                    "- newanswers: Generate new answers (use with caution!)",
                ]
            )
        )


if __name__ == "__main__":
    main(len(sys.argv) > 1 and sys.argv[1] or "")
