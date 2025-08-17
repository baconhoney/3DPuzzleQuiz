from openpyxl import load_workbook
from random import sample, shuffle
from sys import argv
import datetime
import locale
locale.setlocale(locale.LC_ALL, 'hu_HU')

def clean(s: str) -> str:
    return str(s or "").strip(" \t\n")


testcount = 8


print("Usage: python generator.py [-s]\n-s: reshuffle numbers\n")
templatedir = "Minták/"
datanames = ["Név", "Ország", "Város", "Angol Név", "Angol Ország", "Angol Város", "id", "Doboz", "Szám"]
doreshuffle = len(argv) == 2 and argv[1] == "-s"

data = []
masterwb = load_workbook("MesterLista.xlsx")
masterws = masterwb["Lista"]
header = [c.value for c in masterws[1]]
data = [dict(zip(header, [clean(x) for x in r])) for r in masterws.iter_rows(min_row=2, max_row=masterws.max_row, values_only=True)]


if doreshuffle:
    print("Shuffling numbers")
    shuffle(data)
    for i in range(1, len(data) + 1): data[i-1]["Szám"] = i

print("Generating master lists...")
wb_lists = load_workbook(templatedir + "MintaListák.xlsx", read_only=False)
ws_name = wb_lists.copy_worksheet(wb_lists.active)
ws_id = wb_lists.copy_worksheet(wb_lists.active)
ws_box = wb_lists.copy_worksheet(wb_lists.active)
ws_number = wb_lists.copy_worksheet(wb_lists.active)
wb_lists.remove(wb_lists.active)
ws_name.title = "Név"
ws_id.title = "ID"
ws_box.title = "Doboz"
ws_number.title = "Szám"
splitpoint = int((len(data)+1) / 2)
print("Splitpoint:", splitpoint)
for rowindex, row in enumerate(sorted(data, key=lambda x: locale.strxfrm(x["Név"]))):
    for c, d in enumerate(datanames):
        #ws_name.cell((rowindex % splitpoint) + 3, c + (rowindex >= splitpoint and 10 or 1), row[d] or "")
        ws_name.cell(rowindex + 3, c + 1, row[d] or "")
        if doreshuffle: masterws.cell(rowindex + 2, c + 1, row[header[c]])
for rowindex, row in enumerate(sorted(data, key=lambda x: x["id"])):
    for c, d in enumerate(datanames):
        #ws_id.cell((rowindex % splitpoint) + 3, c + (rowindex >= splitpoint and 10 or 1), row[d] or "")
        ws_id.cell(rowindex + 3, c + 1, row[d] or "")
for rowindex, row in enumerate(sorted(data, key=lambda x: x["Doboz"] or "0000" + x["id"])):
    for c, d in enumerate(datanames):
        #ws_box.cell((rowindex % splitpoint) + 3, c + (rowindex >= splitpoint and 10 or 1), row[d] or "")
        ws_box.cell(rowindex + 3, c + 1, row[d] or "")
for rowindex, row in enumerate(sorted(data, key=lambda x: int(x["Szám"]))):
    for c, d in enumerate(datanames):
        #ws_number.cell((rowindex % splitpoint) + 3, c + (rowindex >= splitpoint and 10 or 1), row[d] or "")
        ws_number.cell(rowindex + 3, c + 1, row[d] or "")
wb_lists.save("Listák.xlsx")
if doreshuffle: masterwb.save("MesterLista.xlsx")


print("Generating test100...")
quarterpoint = int((len(data)+3) / 4)
print("Quarterpoint:", quarterpoint)
wb_test100 = load_workbook(templatedir + "Minta100asTeszt.xlsx", read_only=False)
ws_test100_hu = wb_test100["Minta100asTeszt HU"]
ws_test100_en = wb_test100["Minta100asTeszt EN"]
ws_test100sols = wb_test100["MintaMegoldókulcs"]
ws_test100_hu.title = "100as Teszt HU"
ws_test100_en.title = "100as Teszt EN"
ws_test100sols.title = "100as Megoldókulcsok"
ws_test100_hu["A1"] = "100-as Teszt"
ws_test100_hu["A2"] = f"Kutatók éjszakája {datetime.datetime.now().year}"
ws_test100_hu["D1"] = "100-as Teszt"
ws_test100_hu["D2"] = f"Kutatók éjszakája {datetime.datetime.now().year}"
ws_test100_hu["G1"] = "100-as Teszt"
ws_test100_hu["G2"] = f"Kutatók éjszakája {datetime.datetime.now().year}"
ws_test100_hu["J1"] = "100-as Teszt"
ws_test100_hu["J2"] = f"Kutatók éjszakája {datetime.datetime.now().year}"
ws_test100_en["A1"] = "100-as Teszt"
ws_test100_en["A2"] = f"Kutatók éjszakája {datetime.datetime.now().year}"
ws_test100_en["D1"] = "100-as Teszt"
ws_test100_en["D2"] = f"Kutatók éjszakája {datetime.datetime.now().year}"
ws_test100_en["G1"] = "100-as Teszt"
ws_test100_en["G2"] = f"Kutatók éjszakája {datetime.datetime.now().year}"
ws_test100_en["J1"] = "100-as Teszt"
ws_test100_en["J2"] = f"Kutatók éjszakája {datetime.datetime.now().year}"

for rowindex, row in enumerate(sorted(data, key=lambda x: locale.strxfrm(x["Név"]))):
    ws_test100_hu.cell((rowindex % quarterpoint) + 6, int(rowindex / quarterpoint) * 3 + 1, row["Név"])
    ws_test100_hu.cell((rowindex % quarterpoint) + 6, int(rowindex / quarterpoint) * 3 + 2, row["Ország"] + ("" if row["Város"] == "[elsüllyedt]" or row["Város"] == "-" else ", " + row["Város"]))
    ws_test100sols.cell((rowindex % quarterpoint) + 3, int(rowindex / quarterpoint) + 1, int(row["Szám"]))
for rowindex, row in enumerate(sorted(data, key=lambda x: locale.strxfrm(x["Angol Név"]))):
    ws_test100_en.cell((rowindex % quarterpoint) + 6, int(rowindex / quarterpoint) * 3 + 1,  row["Angol Név"])
    ws_test100_en.cell((rowindex % quarterpoint) + 6, int(rowindex / quarterpoint) * 3 + 2, ("" if row["Város"] == "[elsüllyedt]" or row["Város"] == "-" else row["Angol Város"] + ", ") + row["Angol Ország"])
    ws_test100sols.cell((rowindex % quarterpoint) + 3, int(rowindex / quarterpoint) + 5, int(row["Szám"]))
wb_test100.save("100asTeszt.xlsx")


wbTest = load_workbook(templatedir + "MintaTesztek.xlsx", read_only=False)
wbSol = load_workbook(templatedir + "MintaMegoldókulcsok.xlsx", read_only=False)
wsTest_hu = wbTest["MintaHU"]
wsTest_en = wbTest["MintaEN"]
wsSol_hu = wbSol.active
wsSol_en = wbSol.copy_worksheet(wsSol_hu)
wsSol_hu.title = "Kulcs HU"
wsSol_en.title = "Kulcs EN"
for testnum in range(1, testcount+1):
    print(f"Generating test {testnum}")
    testdata = sample(data, 20)
    hu = wbTest.copy_worksheet(wsTest_hu)
    en = wbTest.copy_worksheet(wsTest_en)
    hu.title = f"Teszt {testnum:02} HU"
    en.title = f"Teszt {testnum:02} EN"
    hu["A1"] = f"Teszt {testnum:02}"
    en["A1"] = f"Teszt {testnum:02}"
    hu["A2"] = f"Kutatók éjszakája {datetime.datetime.now().year}"
    en["A2"] = f"Kutatók éjszakája {datetime.datetime.now().year}"
    wsSol_hu.cell(1, testnum, f"Teszt {testnum:02} HU")
    wsSol_en.cell(1, testnum, f"Teszt {testnum:02} EN")
    wsSol_hu.cell(22, testnum, f"Teszt {testnum:02} HU")
    wsSol_en.cell(22, testnum, f"Teszt {testnum:02} EN")
    for rowindex, row in enumerate(sorted(testdata, key=lambda x: locale.strxfrm(x["Név"]))):
        hu.cell(rowindex + 6, 1, row["Név"])
        hu.cell(rowindex + 6, 2, row["Ország"] + ("" if row["Város"] == "[elsüllyedt]" or row["Város"] == "-" else ", " + row["Város"]))
        wsSol_hu.cell(rowindex+2, testnum, int(row["Szám"]))
    for rowindex, row in enumerate(sorted(testdata, key=lambda x: locale.strxfrm(x["Angol Név"]))):
        en.cell(rowindex + 6, 1, row["Angol Név"])
        en.cell(rowindex + 6, 2, ("" if row["Város"] == "[elsüllyedt]" or row["Város"] == "-" else row["Angol Város"] + ", ") + row["Angol Ország"])
        wsSol_en.cell(rowindex+2, testnum, int(row["Szám"]))
wbTest.remove(wsTest_hu)
wbTest.remove(wsTest_en)
wbTest.save("Tesztek.xlsx")
wbSol.save("Megoldókulcsok.xlsx")

